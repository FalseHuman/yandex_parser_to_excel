import requests
import os
import openpyxl
from io import BytesIO
import time
from selenium import webdriver
from selenium.webdriver.chrome.options import Options


def create_driver():
    options = Options()
    options.add_argument(
        f"user-data-dir=YOUR_PATH_PROFILE_CHROME")
    options.add_argument(
        "user-agent=Mozilla/5.0 (Windows NT 6.1; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/84.0.4147.125 Safari/537.36")
    options.add_argument("--headless")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--v=99")
    options.add_argument("--no-sandbox")
    driver = webdriver.Chrome(
        executable_path='YOUR_PATH_CHROMEDRIVER', chrome_options=options)
    return driver

def min_price(prices):
    min_price = []
    for price in prices:
        if '₽' in price:
            price = price.replace('₽', '')
            price = price.split()
            price = ''.join(price)
            min_price.append(int(price))
    return min(min_price)

def original_photo_size(link):
    # Замена значений кропа на orig
    if link != 'https://avatars.mds.yandex.net/get-yapic/0/0-0/islands-retina-50':
        link = link.split('/')
        link[-1] = link[-1].replace(link[-1], 'orig')
        # Формирование новой ссылки
        new_link = ''
        for index_elem in link:
            new_link += index_elem + '/'
        return new_link
    else:
        return link


def clear_text(text):
    garbages = ['Задать вопрос о товаре', 'Подробнее', 'Все товары',
                'Перед покупкой уточняйте характеристики и комплектацию у продавца.', 'Пожаловаться на товар', 'Внешний вид товаров и/или упаковки может быть изменён изготовителем и отличаться от изображенных на Яндекс.Маркете.', 'Есть противопоказания, посоветуйтесь с врачом.']

    for garbage in garbages:
        if garbage in text:
            text = text.replace(garbage, '')
    return text


def feedback_yandex_market(link):
    driver = create_driver()
    driver.get(link)
    time.sleep(10)
    feedback_arr = []
    rating_feedback = {
        'Отличный товар': 5,
        'Хороший товар': 4,
        'Обычный товар': 3,
        'Плохой товар': 2,
        'Ужасный товар': 1
    }

    try:
        feedbacks = driver.find_elements_by_class_name('_13uSY')
        for feedback in feedbacks:
            feedback_dict = {}
            try:
                image_author = feedback.find_element_by_class_name(
                    '_3ZjdE').get_attribute('src')
                image_author = original_photo_size(image_author)
            except:
                image_author = 'Нет аватарки'
            feedback_dict['avatar'] = image_author
            name_user = feedback.find_element_by_class_name('_1mJcZ')
            feedback_dict['name'] = name_user.text
            try:
                image_feedback = feedback.find_element_by_class_name(
                    '_1Tcsj').get_attribute('src')
                image_feedback  = original_photo_size(image_feedback)
            except:
                image_feedback = 'Нет фото у отзыва'
            feedback_dict['photo_feed'] = image_feedback
            #text_feedback = feedback.find_element_by_class_name('_3IXcz')
            try:
                text_feedback = feedback.find_element_by_class_name(
                    '_3IXcz').text
            except:
                text_feedback = 'Нет текста у отзыва'
            feedback_dict['text_feedback'] = text_feedback
            data_feedback = feedback.find_element_by_class_name('kx7am')
            feedback_dict['data_feedback'] = data_feedback.text
            rating = feedback.find_element_by_class_name('pcIgr')
            feedback_dict['rating'] = rating_feedback[rating.text]
            feedback_arr.append(feedback_dict)
        driver.quit()
        return feedback_arr
    except Exception as e:
        print(f'Ошибка парсинга {e}')
        driver.quit()
        return f'error'


def questions_yandex_market(link):
    driver = create_driver()
    driver.get(link)
    time.sleep(10)
    questions_arr = []
    try:
        quetstions = driver.find_elements_by_class_name('_17NwA')
        for queststion in quetstions:
            try:
                image_author = queststion.find_element_by_class_name(
                    '_3ZjdE').get_attribute('src')
                image_author = original_photo_size(image_author)
            except:
                image_author = 'Нет аватарки'
            queststion = queststion.text.split('Ответить')
            queststion_dict = {}
            for comment in range(len(queststion)):
                if comment == 0:
                    # print('Вопрос')
                    comment_split = queststion[0].split('\n')
                    queststion_dict['comment'] = {
                        'user': comment_split[0], 'data': comment_split[1], 'text': comment_split[2]}
                    #print(comment_split[0], comment_split[1], comment_split[2])
                else:
                    # print("Ответ")
                    answer_split = queststion[1].split('\n')
                    if len(answer_split) > 2:
                        queststion_dict['answer'] = {
                            'user': answer_split[2], 'data': answer_split[3], 'text': answer_split[4], 'image': image_author}
                        #print(answer_split[2], answer_split[3], answer_split[4], image_author)
                    else:
                        answer_split = "Нет данных"
                        queststion_dict['answer'] = {
                            'user': answer_split, 'data': answer_split, 'text': answer_split, 'image': answer_split}
            questions_arr.append(queststion_dict)
            # print(questions_arr)
        driver.quit()
        return questions_arr
    except Exception as e:
        print(f'Ошибка парсинга {e}')
        driver.quit()
        return f'error'


def parser_yandex_market(link):
    driver = create_driver()
    driver.get(link)
    time.sleep(10)
    dict_market = {}
    try:
        print(
            'Название, цена, рейтинг, характеристики рядом с заголовком, коротко о товаре')
        name = driver.find_element_by_class_name('_2OAAC').text
        price = (driver.find_element_by_class_name('KnVez'))
        price = (price.text).split('\n')
        pars_price = min_price(price)
        try:
            rating = driver.find_element_by_class_name('_1NfPD').text
        except:
            rating = 'Нет данных'
        try:
            characteristic_h1 = driver.find_element_by_class_name(
                '_2o3uY').text  # TODO разбраться что за хрень!
        except:
            characteristic_h1 = 'Нет данных'
        try:
            user_change = driver.find_element_by_class_name(
                '_2lqOc').text
        except:
            user_change = 'Нет данных'
        try:
            characteristic_low = driver.find_element_by_class_name(
                'Ksay3').text
        except:
            characteristic_low = driver.find_element_by_class_name(
                '_2ZGXi')
            characteristic_low = characteristic_low.find_element_by_class_name(
                'cia-cs').text
        dict_market['name'] = name
        dict_market['price'] = pars_price
        dict_market['rating'] = rating
        dict_market['user_change'] = user_change
        dict_market['characteristic_h1'] = characteristic_h1
        dict_market['characteristic_low'] = characteristic_low
        print('Картинки')
        try:
            images_button = driver.find_element_by_class_name(
                '_19EZn').click()
        except:
            images_button = driver.find_element_by_class_name(
                '_2ljqQ').click()
        else:
            pass
        time.sleep(2)
        div_images = driver.find_element_by_class_name('_2WKLY')
        images = div_images.find_elements_by_tag_name('img')
        arr_images = []
        for image in images:
            image = original_photo_size(image.get_attribute('src'))
            if image not in arr_images:
                arr_images.append(image)
        # Закрытие картинок
        driver.find_element_by_class_name('_2TKD4').click()
        # print(arr_images)
        dict_market['images'] = arr_images
        print('-------------')

        # Характеристика товара
        print('Характеристика')
        spec_button = driver.find_element_by_class_name(
            '_3mNWJ').click()
        time.sleep(10)
        characteristic_product = driver.find_element_by_id(
            'product-specs')
        #characteristic_product = clear_text(characteristic_product)
        # print la3zd
        try:
            info_characteristic = characteristic_product.find_element_by_class_name('_21d7b').text # Описание
            #print(info_characteristic)
            h2_table = characteristic_product.find_element_by_class_name('Y7zwR').text
            #print(h2_table)
            divs_table = characteristic_product.find_elements_by_class_name('la3zd')
            div_table_string = ''
            for div_table in divs_table:
                h3_table = div_table.find_element_by_class_name('_1yUJ7').text
                #print(h3_table)
                text = ''
                for span in div_table.find_elements_by_class_name('sZB0N'):
                    new_span = span.text.replace('\n', ' ') + '\n'
                    text += new_span
                    """ print(span.text.replace('\n', ' '))
                    print('-------------') """

                div_table_string += h3_table + '\n' + text
            characteristic_product = info_characteristic + '\n' + h2_table +'\n' + div_table_string
            #print(characteristic_product)
        except:
            characteristic_product = clear_text(characteristic_product)
        dict_market['characteristic_product'] = characteristic_product
        # print(characteristic_product)
        print('-------------')
        # Вопросы к товару _23Bz1
        button = driver.find_element_by_class_name('_1SqIf')
        link_button = button.get_attribute('href')
        # print(number)
        button.click()
        time.sleep(10)
        queststion_page = driver.find_elements_by_class_name('_2prNU')
        if len(queststion_page) == 0:
            print(link_button)
            dict_market['questions'] = [link_button]
        else:
            queststion_link = []
            for i in range(len(queststion_page)):
                if i == 0 or i == 1 or i == 2:
                    print(queststion_page[i].get_attribute('href'))
                    if queststion_page[i].get_attribute('href') not in queststion_link:
                        queststion_link.append(queststion_page[i].get_attribute('href'))
                else:
                    pass
            dict_market['questions'] = queststion_link
        # Отзывы к товару сбор ссылок
        feed_bck_btn = driver.find_element_by_class_name('_2J5l3')
        link_feed_bck_btn = feed_bck_btn.get_attribute('href')
        feed_bck_btn.click()
        time.sleep(10)
        try:
            dict_market['text_feedback'] = driver.find_element_by_class_name(
                '_1yQSc').text
        except:
            dict_market['text_feedback'] = 'Нет отзыва от Яндекса'
        feedback_page = driver.find_elements_by_class_name('_2prNU')
        if len(feedback_page) == 0:
            print(link_feed_bck_btn)
            dict_market['feedback'] = [link_feed_bck_btn]
        else:
            feedbck_link = []
            for i in range(len(feedback_page)):
                if i == 0 or i == 1 or i == 2:
                    print(feedback_page[i].get_attribute('href'))
                    if feedback_page[i].get_attribute('href') not in feedbck_link:
                        feedbck_link.append(feedback_page[i].get_attribute('href'))
                else:
                    pass
            dict_market['feedback'] = feedbck_link
        driver.quit()
        return dict_market
    except Exception as e:
        print(f'Ошибка парсинга {e}')
        driver.quit()
        return 'error'


def start_parser(chat_id, link):
    token = os.environ.get('token')
    quetstion_func = parser_yandex_market(link)
    if quetstion_func == 'error':
        text = 'Ошибка при парсинге, повторите попытку'
        requests.post(
            f'https://api.telegram.org/bot{token}/sendMessage?chat_id={chat_id}&text={text}')
    else:
        wb = openpyxl.Workbook()
        # Основная информация
        main_page = wb.create_sheet(
            title='Основная информация', index=0)
        sheet_main_page = wb['Основная информация']
        field_name_main_page = ['Название товара', 'Цена', 'Рейтинг',
                                'Выбор покупателей', 'Характеристики, которые пользователи отметили в отзывах', 'Краткая характеристика', 'Характеристика', 'Отзыв Яндекса']

        photo_field_main_page = ['№','Фото']
        sheet_main_page.append(field_name_main_page)

        sheet_main_page.append([quetstion_func['name'], quetstion_func['price'], quetstion_func['rating'],
                                quetstion_func['user_change'], quetstion_func['characteristic_h1'], quetstion_func['characteristic_low'], quetstion_func['characteristic_product'], quetstion_func['text_feedback']])
        sheet_main_page.append(['', '', '', '', '','', '',''])
        sheet_main_page.append(photo_field_main_page)
        photo_count = 0
        for image in quetstion_func['images']:
            photo_count +=1 
            sheet_main_page.append([photo_count, image])
        # Страница вопросов
        question_page = wb.create_sheet(
            title='Вопросы', index=0)
        sheet_question_page = wb['Вопросы']
        field_name_question_page = ['Автор вопроса', 'Дата комментария', 'Текст', 'Автор ответа',
                                    'Дата ответа', 'Текст', 'Аватарка автора ответа']
        sheet_question_page.append(field_name_question_page)
        for i in quetstion_func['questions']:
            if i != 'error':
                arr_questions = questions_yandex_market(i)
                for index in arr_questions:
                    sheet_question_page.append([index['comment']['user'], index['comment']['data'], index['comment']['text'],
                                                index['answer']['user'], index['answer']['data'], index['answer']['text'], index['answer']['image']])
        feedback_page = wb.create_sheet(
            title='Отзывы', index=0)
        sheet_feedbck_page = wb['Отзывы']
        field_name_feedbck_page = ['Авaтарка', 'Имя пользователя', 'Фото отзыва', 'Текст',
                                   'Дата', 'Рейтинг']
        sheet_feedbck_page.append(field_name_feedbck_page)
        for i in quetstion_func['feedback']:
            if i != 'error':
                arr_feedback = feedback_yandex_market(i)
                for index in arr_feedback:
                    sheet_feedbck_page.append([index['avatar'], index['name'], index['photo_feed'],
                                               index['text_feedback'], index['data_feedback'], index['rating']])
        filename = quetstion_func['name']
        excelfile = f'{filename}.xlsx'
        if '/' in excelfile:
            excelfile = excelfile.replace('/', ' ')
        wb.save(excelfile)
        status = requests.post(
            f'https://api.telegram.org/bot{token}/sendDocument?chat_id={chat_id}', files={'document': open(excelfile, 'rb')})
        print(status)
#start_parser('741541899', 'https://market.yandex.ru/product--nabor-iz-10-par-noskov-moscowsocksclub-m20-miks-iskusstvo-razmer-27-41-43/1439308463?glfilter=14474426%3AMjcgKDQxLTQzKQ_101447518663&glfilter=14871214%3A15098577_101447518663&cpc=zHewvSNgIjPtJC9iD60JL3m43OcnhUqQ-g6C4kDb05rWnvIKBybbR6unOSuLlBsL9g15fnalFtreycesyu-TgSxkTsBYWCoNIxS5QvFMzLDWg2VG_56D7oYSPcsjO6rlKiAnV7kSCqmriKKOOtdM0Vd8TVPSkApivEnqem6eGihq0VwesfusPWNxg1F5Tq3i&sku=101447518663&offerid=7JFFcigWlvMzkapCVGazig&cpa=1')
